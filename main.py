import os
import re
import io
import shutil
import tempfile
import json
from typing import List, Tuple, Dict, Any

from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from openpyxl import load_workbook
from openai import OpenAI

app = FastAPI()

# OpenAI client (clé via variable d'env OPENAI_API_KEY)
client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

# Grec moderne + polytonique
GREEK_RE = re.compile(r"[\u0370-\u03FF\u1F00-\u1FFF]")


def is_greek_text(v) -> bool:
    return isinstance(v, str) and bool(GREEK_RE.search(v))


def is_formula_cell(cell) -> bool:
    # openpyxl: data_type "f" = formula
    if getattr(cell, "data_type", None) == "f":
        return True
    v = cell.value
    return isinstance(v, str) and v.lstrip().startswith("=")


def cleanup_files(*paths: str) -> None:
    for p in paths:
        try:
            if p and os.path.exists(p):
                os.remove(p)
        except Exception:
            pass


def _strip_code_fences(s: str) -> str:
    """
    Defensive: parfois un modèle renvoie ```json ... ```
    On enlève les fences si présentes.
    """
    s = (s or "").strip()
    if s.startswith("```"):
        # remove first fence line
        s = s.split("\n", 1)[1] if "\n" in s else ""
        # remove last fence
        if "```" in s:
            s = s.rsplit("```", 1)[0]
    return s.strip()


def translate_batch_json(texts: List[str], model: str) -> List[str]:
    """
    Envoie une LISTE JSON [{id, text}, ...] et attend une LISTE JSON
    [{id, translation}, ...] — même taille, mêmes ids.
    => robuste aux retours à la ligne dans les cellules.
    """
    payload = [{"id": i, "text": t} for i, t in enumerate(texts)]

    resp = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": (
                    "You translate Greek financial spreadsheet labels into clear English.\n"
                    "Rules:\n"
                    "- Do NOT change numbers, dates, tickers, abbreviations, or punctuation.\n"
                    "- Keep terminology consistent (Revenue, EBITDA, Working Capital).\n"
                    "- Output MUST be valid JSON only.\n"
                    "- Output MUST be a JSON array of objects: "
                    "[{\"id\": <int>, \"translation\": <string>}]\n"
                    "- The array MUST contain exactly the same ids as the input, "
                    "no extra items, no missing items.\n"
                    "- Do NOT add explanations, no markdown, no code fences.\n"
                    "- IMPORTANT: Do NOT introduce new newline characters. "
                    "If the source text contains newlines, keep them as-is.\n"
                ),
            },
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ],
    )

    raw_out = _strip_code_fences((resp.output_text or "").strip())

    try:
        out_json = json.loads(raw_out)
    except Exception:
        # on renvoie un bout pour debug sans exploser les logs
        snippet = raw_out[:300].replace("\n", "\\n")
        raise HTTPException(status_code=500, detail=f"Model did not return valid JSON. Got: {snippet}")

    if not isinstance(out_json, list) or len(out_json) != len(payload):
        raise HTTPException(
            status_code=500,
            detail=f"Translation count mismatch: expected {len(payload)}, got {len(out_json) if isinstance(out_json, list) else 'non-list'}",
        )

    id2t: Dict[int, str] = {}
    for obj in out_json:
        if not isinstance(obj, dict) or "id" not in obj or "translation" not in obj:
            raise HTTPException(status_code=500, detail="Invalid JSON schema from model output")
        try:
            i = int(obj["id"])
        except Exception:
            raise HTTPException(status_code=500, detail="Invalid 'id' type in model output")
        id2t[i] = str(obj["translation"])

    if len(id2t) != len(payload):
        raise HTTPException(status_code=500, detail="Duplicate/missing ids in model output")

    # Recompose dans l’ordre original (0..n-1)
    out = []
    for i in range(len(payload)):
        if i not in id2t:
            raise HTTPException(status_code=500, detail=f"Missing id {i} in model output")
        out.append(id2t[i])

    return out


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/translate")
async def translate(
    file: UploadFile = File(...),
    model: str = Query(default="gpt-4.1-mini"),
    batch_size: int = Query(default=40, ge=10, le=120),
):
    # ---- Validate filename/ext ----
    original_name = file.filename or "input.xlsx"
    fname = original_name.lower()

    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="Upload a .xlsx or .xlsm file")

    is_xlsm = fname.endswith(".xlsm")
    out_ext = "xlsm" if is_xlsm else "xlsx"
    base = original_name.rsplit(".", 1)[0]
    out_name = f"{base}_EN.{out_ext}"

    # ---- Save upload to disk (avoid raw bytes in RAM) ----
    in_suffix = ".xlsm" if is_xlsm else ".xlsx"
    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=in_suffix)
    tmp_in_path = tmp_in.name
    tmp_in.close()

    try:
        with open(tmp_in_path, "wb") as f:
            file.file.seek(0)
            shutil.copyfileobj(file.file, f)
    except Exception as e:
        cleanup_files(tmp_in_path)
        raise HTTPException(status_code=400, detail=f"Failed to store upload: {e}")

    # ---- Load workbook ----
    try:
        wb = load_workbook(
            filename=tmp_in_path,
            data_only=False,
            keep_vba=is_xlsm,
        )
    except Exception as e:
        cleanup_files(tmp_in_path)
        raise HTTPException(status_code=400, detail=f"Cannot open workbook: {e}")

    # ---- Translation loop (batch on the fly) ----
    cache: Dict[str, str] = {}

    pending: List[Tuple[object, str]] = []  # (cell, original_text)
    pending_texts: List[str] = []

    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    if is_formula_cell(cell):
                        continue
                    if not is_greek_text(v):
                        continue

                    # Cache hit
                    if v in cache:
                        cell.value = cache[v]
                        continue

                    pending.append((cell, v))
                    pending_texts.append(v)

                    if len(pending_texts) >= batch_size:
                        translated = translate_batch_json(pending_texts, model=model)
                        for (c, orig), tr in zip(pending, translated):
                            cache[orig] = tr
                            c.value = tr
                        pending.clear()
                        pending_texts.clear()

        # flush remaining
        if pending_texts:
            translated = translate_batch_json(pending_texts, model=model)
            for (c, orig), tr in zip(pending, translated):
                cache[orig] = tr
                c.value = tr

    except HTTPException:
        cleanup_files(tmp_in_path)
        raise
    except Exception as e:
        cleanup_files(tmp_in_path)
        raise HTTPException(status_code=500, detail=f"Processing failed: {e}")

    # ---- Save output to disk ----
    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=f".{out_ext}")
    tmp_out_path = tmp_out.name
    tmp_out.close()

    try:
        wb.save(tmp_out_path)
    except Exception as e:
        cleanup_files(tmp_in_path, tmp_out_path)
        raise HTTPException(status_code=500, detail=f"Failed to save output: {e}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    cleanup_files(tmp_in_path)

    return FileResponse(
        path=tmp_out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_name,
        background=BackgroundTask(cleanup_files, tmp_out_path),
    )
